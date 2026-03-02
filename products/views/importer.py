
import uuid
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone

from openpyxl import Workbook, load_workbook

from products.models import Item, Partner, MoveType, StockBalance, StockMove
from products.views.inventory import _role_filter_kwargs

HEADER_FIELDS = [
    ("warehouse", "仓库"),
    ("item", "物品"),
    ("quantity", "数量"),
    ("reference", "单号/来源"),
    ("partner", "合作方"),
    ("note", "备注"),
]
HEADER_LOOKUP = {label: key for key, label in HEADER_FIELDS}
HEADER_LABELS = {key: label for key, label in HEADER_FIELDS}
REQUIRED_FIELDS = {"warehouse", "item", "quantity"}
MAX_ROWS = getattr(settings, "STOCK_IMPORT_MAX_ROWS", 500)
SESSION_PREFIX = "stock_import_"
SESSION_TTL_SECONDS = getattr(settings, "STOCK_IMPORT_SESSION_TTL", 30 * 60)
ACTION_CHOICES = [
    (MoveType.INBOUND, "批量入库"),
    (MoveType.OUTBOUND, "批量出库"),
]
ACTION_TYPES = {choice for choice, _ in ACTION_CHOICES}
ACTION_LABELS = dict(ACTION_CHOICES)


def _session_key(session_id: str) -> str:
    return f"{SESSION_PREFIX}{session_id}"


def _clean_text(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _store_session_data(request, session_id: str, data: dict) -> None:
    request.session[_session_key(session_id)] = data
    request.session.modified = True


def _load_session_data(request, session_id: str):
    data = request.session.get(_session_key(session_id))
    if not data:
        return None
    created_ts = data.get("created_ts")
    if not created_ts:
        return data
    if timezone.now().timestamp() - created_ts > SESSION_TTL_SECONDS:
        request.session.pop(_session_key(session_id), None)
        request.session.modified = True
        return None
    return data


def _build_lookups(role_context):
    warehouses = list(role_context["warehouse"].order_by("name"))
    warehouse_lookup = {w.name.strip().lower(): w for w in warehouses}
    if not warehouse_lookup:
        raise ValueError("当前账号没有可操作的仓库")
    allowed_ids = [w.id for w in warehouses]
    item_lookup = {}
    for item in Item.objects.filter(warehouse_id__in=allowed_ids, is_active=True).select_related("unit"):
        key = (item.warehouse_id, item.name.strip().lower())
        item_lookup[key] = item
    partners = {p.name.strip().lower(): p for p in Partner.objects.filter(is_active=True)}
    return warehouse_lookup, item_lookup, partners


def _parse_excel(uploaded_file, role_context):
    file_bytes = uploaded_file.read()
    if not file_bytes:
        raise ValueError("上传的文件为空")
    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:  # pragma: no cover - openpyxl raises many types
        raise ValueError("无法读取 Excel 文件，请确认格式为 .xlsx") from exc

    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValueError("Excel 中缺少表头行")

    column_map = {}
    for idx, raw_header in enumerate(header_row):
        header_label = _clean_text(raw_header)
        mapped = HEADER_LOOKUP.get(header_label)
        if mapped and mapped not in column_map:
            column_map[mapped] = idx

    missing = REQUIRED_FIELDS - set(column_map.keys())
    if missing:
        raise ValueError("表头缺少必要列：" + ", ".join(sorted(HEADER_LABELS[field] for field in missing)))

    warehouse_lookup, item_lookup, partner_lookup = _build_lookups(role_context)

    rows = []
    for row_idx, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row_data = {}
        for field, col_idx in column_map.items():
            row_data[field] = values[col_idx] if values and col_idx < len(values) else None

        warehouse_name = _clean_text(row_data.get("warehouse"))
        item_name = _clean_text(row_data.get("item"))
        quantity_raw = row_data.get("quantity")
        reference = _clean_text(row_data.get("reference"))
        partner_name = _clean_text(row_data.get("partner"))
        note = _clean_text(row_data.get("note"))

        if not any([warehouse_name, item_name, _clean_text(quantity_raw)]):
            continue

        error = None
        warehouse_obj = None
        item_obj = None
        partner_obj = None
        quantity_value = ""

        if not warehouse_name:
            error = "仓库为空"
        else:
            warehouse_obj = warehouse_lookup.get(warehouse_name.lower())
            if not warehouse_obj:
                error = "仓库不存在或无权限"

        if not error and not item_name:
            error = "物品名称为空"
        elif not error:
            item_obj = item_lookup.get((warehouse_obj.id, item_name.lower())) if warehouse_obj else None
            if not item_obj:
                error = "物品不存在或未启用"

        if not error:
            if quantity_raw in (None, ""):
                error = "数量为空"
            else:
                try:
                    qty_decimal = Decimal(str(quantity_raw))
                except (InvalidOperation, TypeError):
                    error = "数量需为数字"
                else:
                    if qty_decimal <= 0:
                        error = "数量必须大于 0"
                    elif qty_decimal != qty_decimal.to_integral_value():
                        error = "数量必须是整数"
                    else:
                        quantity_value = str(int(qty_decimal))

        if not error and partner_name:
            partner_obj = partner_lookup.get(partner_name.lower())
            if not partner_obj:
                error = "合作方不存在或已停用"

        rows.append({
            "index": row_idx,
            "warehouse_name": warehouse_obj.name if warehouse_obj else warehouse_name,
            "warehouse_id": warehouse_obj.id if warehouse_obj else None,
            "item_name": item_obj.name if item_obj else item_name,
            "item_id": item_obj.id if item_obj else None,
            "unit_name": item_obj.unit.name if item_obj else "",
            "quantity": quantity_value,
            "reference": reference,
            "partner_name": partner_obj.name if partner_obj else partner_name,
            "partner_id": partner_obj.id if partner_obj else None,
            "note": note,
            "error": error,
        })

        if len(rows) > MAX_ROWS:
            raise ValueError(f"单次最多导入 {MAX_ROWS} 条，请拆分文件后再上传")

    if not rows:
        raise ValueError("Excel 中没有任何有效数据")

    return rows

@login_required
def stock_import_start(request):
    if request.method == "POST":
        action_type = request.POST.get("action_type")
        if action_type not in ACTION_TYPES:
            messages.error(request, "请选择入库或出库类型")
            return redirect(reverse("products:stock_import_start"))

        upload = request.FILES.get("file")
        if not upload:
            messages.error(request, "请先选择要上传的 Excel 文件")
            return redirect(reverse("products:stock_import_start"))

        try:
            rows = _parse_excel(upload, _role_filter_kwargs(request.user))
        except ValueError as exc:
            messages.error(request, str(exc))
            return redirect(reverse("products:stock_import_start"))

        session_id = uuid.uuid4().hex
        session_data = {
            "id": session_id,
            "action_type": action_type,
            "filename": upload.name,
            "rows": rows,
            "created_ts": timezone.now().timestamp(),
        }
        _store_session_data(request, session_id, session_data)
        return redirect(reverse("products:stock_import_preview", args=[session_id]))

    return render(request, "products/stock_import_upload.html", {
        "action_choices": ACTION_CHOICES,
        "max_rows": MAX_ROWS,
        "header_fields": [label for _, label in HEADER_FIELDS],
    })


@login_required
def stock_import_preview(request, session_id: str):
    session_data = _load_session_data(request, session_id)
    if not session_data:
        messages.error(request, "导入会话已失效，请重新上传 Excel")
        return redirect(reverse("products:stock_import_start"))

    rows = session_data.get("rows", [])
    has_error = any(row.get("error") for row in rows)
    valid_count = sum(1 for row in rows if not row.get("error"))
    action_type = session_data.get("action_type")
    action_label = ACTION_LABELS.get(action_type, action_type)

    return render(request, "products/stock_import_preview.html", {
        "session_id": session_id,
        "data": session_data,
        "rows": rows,
        "has_error": has_error,
        "valid_count": valid_count,
        "total_count": len(rows),
        "action_label": action_label,
    })


@login_required
def stock_import_confirm(request, session_id: str):
    if request.method != "POST":
        return redirect(reverse("products:stock_import_preview", args=[session_id]))

    session_data = _load_session_data(request, session_id)
    if not session_data:
        messages.error(request, "导入会话已失效，请重新上传 Excel")
        return redirect(reverse("products:stock_import_start"))

    rows = session_data.get("rows", [])
    if not rows:
        messages.error(request, "没有可导入的数据")
        return redirect(reverse("products:stock_import_start"))

    if any(row.get("error") for row in rows):
        messages.error(request, "仍有错误行，无法导入。请返回预览页修正后再试")
        return redirect(reverse("products:stock_import_preview", args=[session_id]))

    action_type = session_data.get("action_type")
    if action_type not in ACTION_TYPES:
        messages.error(request, "未知的导入类型")
        return redirect(reverse("products:stock_import_start"))

    role_context = _role_filter_kwargs(request.user)
    allowed_ids = set(role_context["warehouse"].values_list("id", flat=True))
    if not allowed_ids:
        messages.error(request, "当前账号没有可操作的仓库")
        return redirect(reverse("products:stock_import_start"))

    outbound_requirements = defaultdict(int)
    item_pairs = set()
    for row in rows:
        warehouse_id = row.get("warehouse_id")
        item_id = row.get("item_id")
        quantity = int(row["quantity"])
        if not warehouse_id or warehouse_id not in allowed_ids:
            messages.error(request, "检测到无权限的仓库记录，请重新上传")
            return redirect(reverse("products:stock_import_start"))
        if not item_id:
            messages.error(request, "检测到无效的数据，请重新上传")
            return redirect(reverse("products:stock_import_start"))
        item_pairs.add((warehouse_id, item_id))
        if action_type == MoveType.OUTBOUND:
            outbound_requirements[(warehouse_id, item_id)] += quantity

    if outbound_requirements:
        warehouse_ids = sorted({wid for wid, _ in item_pairs})
        item_ids = sorted({iid for _, iid in item_pairs})
        if warehouse_ids and item_ids:
            balances = {
                (bal.warehouse_id, bal.item_id): bal.on_hand
                for bal in StockBalance.objects.filter(
                    warehouse_id__in=warehouse_ids,
                    item_id__in=item_ids,
                )
            }
        else:
            balances = {}
        insufficient = []
        for key, need in outbound_requirements.items():
            on_hand = balances.get(key, 0)
            if on_hand < need:
                insufficient.append((key, on_hand, need))
        if insufficient:
            first = insufficient[0]
            messages.error(
                request,
                f"出库失败：库存不足（仓库ID {first[0][0]} 物品ID {first[0][1]} 当前 {first[1]}，需 {first[2]}）。请检查库存后再试",
            )
            return redirect(reverse("products:stock_import_preview", args=[session_id]))

    batch_reference = f"BATCH-{timezone.now().strftime('%Y%m%d%H%M%S')}-{request.user.id}"

    try:
        with transaction.atomic():
            for row in rows:
                quantity = int(row["quantity"])
                qty_value = quantity if action_type == MoveType.INBOUND else -quantity
                StockMove.objects.create(
                    move_type=action_type,
                    warehouse_id=row["warehouse_id"],
                    item_id=row["item_id"],
                    quantity=qty_value,
                    reference=row.get("reference") or batch_reference,
                    note=row.get("note", ""),
                    partner_id=row.get("partner_id"),
                )
    except Exception:
        messages.error(request, "导入失败，请重试或联系管理员")
        return redirect(reverse("products:stock_import_preview", args=[session_id]))

    request.session.pop(_session_key(session_id), None)
    request.session.modified = True

    messages.success(request, f"已成功导入 {len(rows)} 条记录")
    return redirect(reverse("products:inventory_dashboard"))


@login_required
def stock_import_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"
    ws.append([label for _, label in HEADER_FIELDS])
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=stock_import_template.xlsx"
    wb.save(response)
    return response
